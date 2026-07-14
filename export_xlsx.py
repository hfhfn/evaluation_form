"""导出成绩为 Excel（.xlsx）。

两个工作表：
  1. 「评分明细」——每位评分人对每组的逐维度打分、总分、评语（与原 CSV 明细一致）。
  2. 「排名与评语」——所有组的排名 / 各维度均分 / 综合平均分，并在下方按组逐条列出收到的评语。

数据来源统一为 db.get_results() 的返回结构，保证两表口径一致。
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="1A73E8")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SECTION_FONT = Font(bold=True, color="0D3F8F", size=12)
_GROUP_FONT = Font(bold=True, color="0D3F8F")
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, ncols: int, row_idx: int = 1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER


def build_results_xlsx(data: dict) -> bytes:
    criteria = data.get("criteria", [])
    groups = data.get("groups", {})
    ranked = data.get("ranked", [])

    wb = Workbook()

    # ---------------- Sheet1：评分明细 ----------------
    ws1 = wb.active
    ws1.title = "评分明细"
    header1 = (["评分人", "评分人所在组", "被评组号"]
               + [f"{c['label']}(分)" for c in criteria]
               + ["总分", "评语"])
    ws1.append(header1)
    _style_header(ws1, len(header1))

    for g in sorted(groups):
        for s in groups[g]["scores"]:
            # 按快照维度名匹配（criterion_id 会因编辑评分标准而变，名称快照才稳定）
            smap = {cs["criterion_label"]: cs["score"] for cs in s.get("criteria_scores", [])}
            ws1.append(
                [s["scorer_name"], s["scorer_group"], g]
                + [smap.get(c["label"], "") for c in criteria]
                + [s["total_score"], (s.get("comment") or "")]
            )

    ws1.freeze_panes = "A2"
    for i in range(1, len(header1) + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 13
    comment_col = len(header1)
    ws1.column_dimensions[get_column_letter(comment_col)].width = 50
    for r in range(2, ws1.max_row + 1):
        ws1.cell(row=r, column=comment_col).alignment = _WRAP

    # ---------------- Sheet2：排名与评语 ----------------
    ws2 = wb.create_sheet("排名与评语")
    header2 = (["排名", "组号", "收到评分数"]
               + [f"{c['label']}(均分)" for c in criteria]
               + ["综合平均分"])
    ws2.append(header2)
    _style_header(ws2, len(header2))

    for r in ranked:
        g = r["group_number"]
        grp = groups.get(g, {"scores": []})
        dim_avgs = []
        for c in criteria:
            vals = [cs["score"] for s in grp["scores"]
                    for cs in s.get("criteria_scores", []) if cs["criterion_label"] == c["label"]]
            dim_avgs.append(round(sum(vals) / len(vals), 2) if vals else "")
        ws2.append([r["rank"], f"第{g}组", r["score_count"]] + dim_avgs + [r["avg_total"]])

    ws2.freeze_panes = "A2"
    ncols2 = len(header2)
    for i in range(1, ncols2 + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 13

    # 逐条评语区（排名表下方空一行）
    row = ws2.max_row + 2
    ws2.cell(row=row, column=1, value="各组收到的评语（逐条）").font = _SECTION_FONT
    row += 1
    for r in ranked:
        g = r["group_number"]
        grp = groups.get(g, {"scores": []})
        # 收到的评语——匿名，只保留内容，不含评分人姓名
        comments = [(s.get("comment") or "").strip() for s in grp["scores"]]
        comments = [c for c in comments if c]

        title = ws2.cell(
            row=row, column=1,
            value=f"第{g}组 · 综合平均 {r['avg_total']} · 收到 {r['score_count']} 份评分 · 有效评语 {len(comments)} 条",
        )
        title.font = _GROUP_FONT
        row += 1

        if comments:
            for idx, text in enumerate(comments, 1):
                ws2.cell(row=row, column=2, value=idx)
                cell = ws2.cell(row=row, column=3, value=text)
                cell.alignment = _WRAP
                row += 1
        else:
            ws2.cell(row=row, column=2, value="（暂无评语）")
            row += 1
        row += 1  # 组间空行

    # 评语区列宽（作用于整列，兼顾上方排名表）
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 6
    ws2.column_dimensions["C"].width = 60

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
