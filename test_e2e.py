# -*- coding: utf-8 -*-
"""端到端（E2E）回归测试

把整套业务流程从头到尾跑一遍并逐项断言：默认评分标准、班级管理、
学生导入、学生评分、按班级筛选、CSV 导出、评分模板、删除班级清理等。

特点：
  * 使用独立的临时数据库，不会污染你的 evaluation.db
  * 跑完自动清理临时文件
  * 任意断言失败则以非零状态码退出（方便接入 CI）

运行方式：
    python test_e2e.py
"""

import os
import sys
import shutil
import sqlite3
import tempfile

# ---- 1. 在导入应用之前，把数据库指向独立的临时文件 ----
sys.argv = [sys.argv[0]]  # 清掉额外命令行参数，确保 config 的 argparse 取默认值
from config import config  # noqa: E402

_TMP_DIR = tempfile.mkdtemp(prefix="eval_e2e_")
config.SQLITE_PATH = os.path.join(_TMP_DIR, "test.db")

# ---- 2. 导入应用（此时才会在临时库里建表 + 播种默认数据）----
from main import app  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

client = TestClient(app)

# ---- 测试小工具 ----
_passed = 0
_failed = 0


def check(cond, msg):
    global _passed, _failed
    if cond:
        _passed += 1
        print(f"  ✅ {msg}")
    else:
        _failed += 1
        print(f"  ❌ 失败: {msg}")


def section(title):
    print(f"\n=== {title} ===")


def login():
    r = client.post("/api/admin/login", json={"username": "admin", "password": "admin123"})
    assert r.json().get("ok"), "管理员登录失败，无法继续测试"


def db_query(sql, params=()):
    conn = sqlite3.connect(config.SQLITE_PATH)
    conn.row_factory = sqlite3.Row
    try:
        return conn.execute(sql, params).fetchone()
    finally:
        conn.close()


# ============================================================
#  测试用例
# ============================================================
def run():
    section("默认评分标准（来自 答辩评分模板.md）")
    crit = client.get("/api/criteria").json()["criteria"]
    check(len(crit) == 6, f"默认 6 个评分维度（实际 {len(crit)}）")
    check(crit[0]["label"].startswith("需求分析"), "维度1 = 需求分析+技术架构+可行性")
    emo = crit[5]["options"]
    check(max(o["score"] for o in emo) == 3 and min(o["score"] for o in emo) == 0,
          "情感分取值 0~3")
    total_max = sum(max(o["score"] for o in c["options"]) for c in crit)
    check(total_max == 18, f"满分合计 18（实际 {total_max}）")

    login()

    section("班级管理 — 空班级可创建并持久化")
    client.post("/api/classes", json={"name": "晚班一组"})
    classes = {x["name"]: x["count"] for x in client.get("/api/classes").json()["classes"]}
    check("晚班一组" in classes, "新建的空班级出现在列表中")
    check(classes.get("晚班一组") == 0, "空班级人数为 0")

    section("每班独立评分表 — 建班即生成、与全局/他班互相隔离")
    global_ids = [c["id"] for c in client.get("/api/criteria").json()["criteria"]]
    a_crit = client.get("/api/criteria?class_name=晚班一组").json()["criteria"]
    a_ids = [c["id"] for c in a_crit]
    check(len(a_crit) == 6, "新班级自动获得 6 个维度（复制自全局默认）")
    check(set(a_ids).isdisjoint(set(global_ids)), "班级维度 id 与全局默认完全独立")

    client.post("/api/classes", json={"name": "早班一组"})
    b_ids = [c["id"] for c in client.get("/api/criteria?class_name=早班一组").json()["criteria"]]
    check(set(a_ids).isdisjoint(set(b_ids)), "两个班级的维度 id 互相独立")

    # 改晚班一组的第一个维度名，验证不影响他班与全局
    edited = [dict(c) for c in a_crit]
    edited[0] = dict(edited[0], label="【改】需求分析")
    client.post("/api/criteria", json={"class_name": "晚班一组", "criteria": edited})
    a_after = client.get("/api/criteria?class_name=晚班一组").json()["criteria"]
    g_after = client.get("/api/criteria").json()["criteria"]
    b_after = client.get("/api/criteria?class_name=早班一组").json()["criteria"]
    check(a_after[0]["label"] == "【改】需求分析", "晚班一组维度名已更新")
    check(a_after[0]["id"] == a_ids[0], "编辑保留维度 id（历史不串）")
    check(g_after[0]["label"].startswith("需求分析"), "全局默认不受影响")
    check(b_after[0]["label"].startswith("需求分析"), "早班一组不受影响")

    section("学生导入 — 归班 / 跨班同名 / 同班去重")
    r = client.post("/api/students/import",
                    json={"students": [{"name": "阿强", "group": 1}, {"name": "阿珍", "group": 2}],
                          "class_name": "晚班一组"})
    check(r.json()["added"] == 2, "导入 2 名学生到晚班一组")

    r = client.post("/api/students/import",
                    json={"students": [{"name": "阿强", "group": 1}], "class_name": "早班一组"})
    check(r.json()["added"] == 1, "同名学生在不同班级不被误去重")

    r = client.post("/api/students/import",
                    json={"students": [{"name": "阿强", "group": 1}], "class_name": "晚班一组"})
    check(r.json()["skipped"] == 1, "同班同名学生正确去重")

    section("学生评分 — 用本班评分表提交（携带班级与评语）")
    a_crit = client.get("/api/criteria?class_name=晚班一组").json()["criteria"]
    sels = [{"criterion_id": c["id"], "option_id": c["options"][0]["id"],
             "score": c["options"][0]["score"]} for c in a_crit]
    r = client.post("/api/scores",
                    json={"scorer_name": "阿强", "scorer_group": 1, "target_group": 2,
                          "selections": sels, "comment": "答辩很精彩", "scorer_class": "晚班一组"})
    check(r.json().get("ok"), "评分提交成功")

    r = client.post("/api/scores",
                    json={"scorer_name": "阿强", "scorer_group": 1, "target_group": 2,
                          "selections": sels, "comment": "", "scorer_class": "晚班一组"})
    check(r.json().get("error") == "您已评过该组，不可重复评分", "重复评分被拒绝")

    bad = [{"criterion_id": a_crit[0]["id"], "option_id": 999999, "score": 9}]
    r = client.post("/api/scores",
                    json={"scorer_name": "阿珍", "scorer_group": 2, "target_group": 1,
                          "selections": bad, "comment": "", "scorer_class": "晚班一组"})
    check(r.json().get("error") == "无效的评分选项", "无效选项被拒绝")

    section("成绩汇总 — 按班级筛选，明细维度名正确")
    res = client.get("/api/results?class_name=晚班一组").json()
    g2 = res["groups"].get("2") or res["groups"].get(2)
    check(g2["score_count"] == 1, "按班级筛选能看到该条评分")
    check(g2["scores"][0]["comment"] == "答辩很精彩", "评语正确入库")
    check(g2["scores"][0]["scorer_class"] == "晚班一组", "scorer_class 正确上报")
    cs = g2["scores"][0]["criteria_scores"]
    check(len(cs) == 6, "明细含 6 个维度分")
    check(any(c["criterion_label"] == "【改】需求分析" for c in cs), "明细维度名取自快照（含改后的名）")

    section("Excel 导出（明细 + 排名与评语两个 sheet）")
    import io as _io
    from openpyxl import load_workbook
    xlsx = client.get("/api/results/export?class_name=晚班一组").content
    wb = load_workbook(_io.BytesIO(xlsx))
    check("评分明细" in wb.sheetnames, "含「评分明细」sheet")
    check("排名与评语" in wb.sheetnames, "含「排名与评语」sheet")
    detail_txt = "\n".join(
        str(c) for row in wb["评分明细"].iter_rows(values_only=True) for c in row if c is not None
    )
    check("评分人" in detail_txt, "明细 sheet 含表头")
    check("答辩很精彩" in detail_txt, "明细 sheet 含评语")
    summary_txt = "\n".join(
        str(c) for row in wb["排名与评语"].iter_rows(values_only=True) for c in row if c is not None
    )
    check("综合平均分" in summary_txt, "排名 sheet 含综合平均分列")
    check("答辩很精彩" in summary_txt, "排名 sheet 逐条列出了评语")

    section("评分标准隔离 — 换标准后旧评分不残留，换回即恢复（数据不丢）")
    # 删掉一个维度（情感分）= 切换到新评分标准（维度集合变了）
    shrunk = a_crit[:-1]
    client.post("/api/criteria", json={"class_name": "晚班一组", "criteria": shrunk})
    check(len(client.get("/api/criteria?class_name=晚班一组").json()["criteria"]) == 5,
          "晚班一组切到 5 维新标准")
    res_new = client.get("/api/results?class_name=晚班一组").json()
    check(sum(g["score_count"] for g in res_new["groups"].values()) == 0,
          "换新标准后旧标准评分不再残留（隔离生效）")
    detail_new = client.get("/api/results/group/2?class_name=晚班一组").json()
    check(len(detail_new["scores"]) == 0, "换新标准后该组明细为空")
    # 换回原 6 维标准（维度名集合一致）
    client.post("/api/criteria", json={"class_name": "晚班一组", "criteria": a_crit})
    res_back = client.get("/api/results?class_name=晚班一组").json()
    g2b = res_back["groups"].get("2") or res_back["groups"].get(2)
    check(g2b["score_count"] == 1, "换回原标准后评分恢复（历史未丢）")
    check(g2b["scores"][0]["comment"] == "答辩很精彩", "恢复的评分内容完整")

    section("清空本班当前标准评分")
    before = client.get("/api/results?class_name=晚班一组").json()
    had = sum(g["score_count"] for g in before.get("groups", {}).values())
    check(had > 0, "清空前该班当前标准有评分")
    r = client.post("/api/scores/clear", json={"class_name": "晚班一组"})
    check(r.json().get("ok") and r.json().get("deleted", 0) > 0, "清空接口返回删除条数")
    after = client.get("/api/results?class_name=晚班一组").json()
    left = sum(g["score_count"] for g in after.get("groups", {}).values())
    check(left == 0, "清空后该班当前标准评分为 0")
    check(client.post("/api/scores/clear", json={"class_name": ""}).json().get("ok") is False,
          "空班级名被拒绝（防误删）")

    section("换标准后可重评（标准感知判重 + 替换旧评分）")
    client.post("/api/classes", json={"name": "X班"})
    client.post("/api/students/import",
                json={"class_name": "X班", "students": [{"name": "小明", "group": 1}, {"name": "小红", "group": 2}]})

    def std(labels):
        return [{"label": l, "options": [{"label": "低", "score": 1}, {"label": "高", "score": 3}]} for l in labels]

    client.post("/api/criteria", json={"class_name": "X班", "criteria": std(["A1", "A2"])})
    crA = client.get("/api/criteria?class_name=X班").json()["criteria"]
    selA = [{"criterion_id": c["id"], "option_id": c["options"][-1]["id"], "score": 3} for c in crA]
    r = client.post("/api/scores", json={"scorer_name": "小明", "scorer_group": 1, "target_group": 2,
                                          "selections": selA, "comment": "", "scorer_class": "X班"})
    check(r.json().get("ok"), "标准A下小明评2组成功")
    my = client.get("/api/scores/my?name=小明").json()["scores"]
    check(any(s["target_group"] == 2 for s in my), "标准A下小明显示已评2组")
    r = client.post("/api/scores", json={"scorer_name": "小明", "scorer_group": 1, "target_group": 2,
                                          "selections": selA, "comment": "", "scorer_class": "X班"})
    check("已评" in (r.json().get("error") or ""), "同标准下重复评2组被拒")

    client.post("/api/criteria", json={"class_name": "X班", "criteria": std(["B1", "B2", "B3"])})
    my2 = client.get("/api/scores/my?name=小明").json()["scores"]
    check(all(s["target_group"] != 2 for s in my2), "换标准B后旧评分不再算已评（可重评）")
    crB = client.get("/api/criteria?class_name=X班").json()["criteria"]
    selB = [{"criterion_id": c["id"], "option_id": c["options"][-1]["id"], "score": 3} for c in crB]
    r = client.post("/api/scores", json={"scorer_name": "小明", "scorer_group": 1, "target_group": 2,
                                          "selections": selB, "comment": "新标准", "scorer_class": "X班"})
    check(r.json().get("ok"), "标准B下小明可重评2组")
    resB = client.get("/api/results?class_name=X班").json()
    g2x = resB["groups"].get("2") or resB["groups"].get(2)
    check(g2x["score_count"] == 1 and g2x["scores"][0]["total_score"] == 9,
          "标准B下只剩 1 条（旧标准评分已被替换）")

    section("评分模板 — 保存 / 列表 / 加载")
    r = client.post("/api/templates", json={"name": "E2E测试模板", "criteria": crit})
    check(r.json().get("ok"), "保存标准为模板")
    check(isinstance(r.json().get("id"), int) and r.json().get("id") > 0, "保存模板返回新 id")
    tpls = client.get("/api/templates").json()["templates"]
    mine = [t for t in tpls if t["name"] == "E2E测试模板"]
    check(len(mine) == 1, "模板出现在列表中")
    loaded = client.get(f"/api/templates/{mine[0]['id']}").json()["criteria"]
    check(len(loaded) == 6, "按 id 加载模板返回 6 个维度")

    section("用模板建班 — 新班直接采用模板评分表")
    client.post("/api/classes", json={"name": "模板班", "template_id": mine[0]["id"]})
    tmpl_crit = client.get("/api/criteria?class_name=模板班").json()["criteria"]
    check(len(tmpl_crit) == 6, "模板班获得 6 个维度")
    check(tmpl_crit[0]["label"].startswith("需求分析"), "模板班维度来自模板")

    section("班级↔当前评分模板 绑定 — 记录 / 读回 / 解绑 / 应用生效")
    # 未绑定时为 None
    check(client.get("/api/class-template?class_name=模板班").json().get("template_id") is None,
          "未绑定时 template_id 为 None")
    # 绑定后读回一致
    client.post("/api/class-template", json={"class_name": "模板班", "template_id": mine[0]["id"]})
    check(client.get("/api/class-template?class_name=模板班").json().get("template_id") == mine[0]["id"],
          "绑定后读回一致")
    # 选模板即应用为该班标准：换成一个 2 维模板后，成绩汇总按新标准隔离（旧评分不再计入）
    two_dim = [{"label": l, "options": [{"label": "低", "score": 1}, {"label": "高", "score": 3}]}
               for l in ["维度甲", "维度乙"]]
    tid2 = client.post("/api/templates", json={"name": "E2E两维模板", "criteria": two_dim}).json()["id"]
    client.post("/api/criteria", json={"class_name": "模板班", "criteria": two_dim})  # 应用=换标准
    client.post("/api/class-template", json={"class_name": "模板班", "template_id": tid2})
    applied = client.get("/api/criteria?class_name=模板班").json()["criteria"]
    check([c["label"] for c in applied] == ["维度甲", "维度乙"], "选模板后该班标准已切换为新模板维度")
    check(client.get("/api/class-template?class_name=模板班").json().get("template_id") == tid2,
          "绑定已更新为新模板")
    # 删除模板后绑定视为解绑（None）
    client.request("DELETE", "/api/templates/%d" % tid2)
    check(client.get("/api/class-template?class_name=模板班").json().get("template_id") is None,
          "绑定模板被删除后 template_id 回落为 None")

    section("模板就地修改（PUT）— 覆盖同一模板")
    up_crit = [{"label": "改后维度", "options": [{"label": "L", "score": 1}, {"label": "H", "score": 3}]}]
    r = client.put("/api/templates/%d" % mine[0]["id"], json={"name": "E2E测试模板改", "criteria": up_crit})
    check(r.json().get("ok"), "PUT 更新模板成功")
    reloaded = client.get("/api/templates/%d" % mine[0]["id"]).json()["criteria"]
    check(len(reloaded) == 1 and reloaded[0]["label"] == "改后维度", "模板内容已就地更新")
    names = [t["name"] for t in client.get("/api/templates").json()["templates"]]
    check("E2E测试模板改" in names, "模板改名生效")

    section("等级顺序持久化 — 保存后顺序与提交顺序一致（不被分值左右）")
    client.post("/api/classes", json={"name": "排序班"})
    ordered = [{"label": "维度O", "sort_order": 0, "options": [
        {"label": "高档", "score": 3, "sort_order": 0},
        {"label": "低档", "score": 1, "sort_order": 1},
        {"label": "中档", "score": 2, "sort_order": 2},
    ]}]
    client.post("/api/criteria", json={"class_name": "排序班", "criteria": ordered})
    got = client.get("/api/criteria?class_name=排序班").json()["criteria"][0]["options"]
    check([o["label"] for o in got] == ["高档", "低档", "中档"],
          "选项顺序按提交的 sort_order 保存（与分值无关）")
    ordered[0]["options"] = [
        {"label": "低档", "score": 1, "sort_order": 0},
        {"label": "高档", "score": 3, "sort_order": 1},
        {"label": "中档", "score": 2, "sort_order": 2},
    ]
    client.post("/api/criteria", json={"class_name": "排序班", "criteria": ordered})
    got2 = client.get("/api/criteria?class_name=排序班").json()["criteria"][0]["options"]
    check([o["label"] for o in got2] == ["低档", "高档", "中档"], "重排后顺序持久生效")

    section("删除班级 — 无孤儿 / 连带删该班评分表 / 只删目标")
    a_crit_ids = [c["id"] for c in client.get("/api/criteria?class_name=晚班一组").json()["criteria"]]
    client.request("DELETE", "/api/classes/晚班一组")
    orphan = db_query(
        "SELECT COUNT(*) AS n FROM score_details WHERE score_id NOT IN (SELECT id FROM scores)"
    )["n"]
    check(orphan == 0, "删除班级后无孤儿 score_details")
    left = db_query("SELECT COUNT(*) AS n FROM students WHERE class_name='晚班一组'")["n"]
    check(left == 0, "晚班一组学生已清空")
    crit_left = db_query("SELECT COUNT(*) AS n FROM criteria WHERE class_name='晚班一组'")["n"]
    check(crit_left == 0, "晚班一组的评分表已删除")
    names = [x["name"] for x in client.get("/api/classes").json()["classes"]]
    check("晚班一组" not in names and "早班一组" in names, "只删目标班级，保留其他班级")
    g_final = client.get("/api/criteria").json()["criteria"]
    check(g_final[0]["label"].startswith("需求分析"), "全局默认评分表始终完好")

    section("权限 — 未登录不可操作")
    anon = TestClient(app)  # 不带登录 cookie
    r = anon.post("/api/classes", json={"name": "黑客班"})
    check(r.status_code == 401, "未登录创建班级被拒（401）")


# ============================================================
#  入口
# ============================================================
if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    print("⏳ 开始端到端测试（使用临时数据库，不影响 evaluation.db）...")
    try:
        run()
    finally:
        client.close()
        shutil.rmtree(_TMP_DIR, ignore_errors=True)

    print(f"\n{'='*46}")
    print(f"  通过 {_passed} 项 / 失败 {_failed} 项")
    print(f"{'='*46}")
    if _failed:
        print("❌ 有用例失败")
        sys.exit(1)
    print("\U0001f389 全部通过")
    sys.exit(0)
